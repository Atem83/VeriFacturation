import sys
import re
from openpyxl import load_workbook
from collections import Counter

"""
Importe la liste des factures,
Créer la liste des doublons et factures manquantes

Auteur : Michaël AUGUSTE
"""

def compte_client_valide(numCompte, debit, credit):
    """Retourne la valeur True si le compte commence par un C
    et que les valeurs debit et credit sont numériques

    Args:
        numCompte (string) : numéro de compte
        debit (int) : montant débit
        credit (int) : montant crédit
    """
    return (isinstance(numCompte, str) and 
            numCompte.startswith('C') and
            isinstance(debit, (int, float)) and 
            isinstance(credit, (int, float)))

def extraire_annee(num_facture):
    """Renvoie l'année à partir du numéro de facture
    Args:
        num_facture (string): numéro de facture
    Returns:
        annee (int) : année de la facture
    """
    
    # Extraire les chiffres avant le premier point
    resultat = re.search(r'\D*(\d+)\.', num_facture)
    
    if resultat:
        annee = int(resultat.group(1))
        return annee
    else:
        return 0

def extraire_mois(num_facture):
    """Renvoie le mois à partir du numéro de facture
    Args:
        num_facture (string): numéro de facture
    Returns:
        mois (int) : mois de la facture au format numérique
    """
    
    # Si ma valeur d'entrée n'est pas au format str, termine la fonction
    if not isinstance(num_facture, str):
        return None
    
    # Extraire les chiffres entre les deux points
    resultat = re.search(r'\.(\d+)\.', num_facture)
    
    if resultat:
        mois = int(resultat.group(1))
        return mois
    else:
        return 0

def extraire_indice(num_facture):
    """Renvoie l'indice mensuel à partir du numéro de facture
    Args:
        num_facture (string) : numéro de facture
    Returns:
        indice (int) : indice mensuel de facturation
    """
    
    # Extraire les chiffres dans la séquence après le dernier point
    # Une lettre A est possible après le point, de manière optionnel
    resultat = re.search(r'\.(\d+)(A)?$', num_facture)
    
    if resultat:
        # group(1) renvoie les chiffres capturés par (\d+)
        # group(2) renvoie None car le groupe (A) optionnel n'est pas présent
        indice = int(resultat.group(1))
        return indice
    else:
        return 0

def extraire_radical(num_facture):
    """Renvoie le radical du numéro de facture
    Args:
        num_facture (string) : numéro de facture
    Returns:
        radical (str) : radical de facturation de type "H2023.02."
    """
    
    # Extraire le radical jusqu'au deuxième point
    resultat = re.search(r'^(.*?)\.[^.]*$', num_facture)
    
    if resultat:
        radical = resultat.group(1) + '.'
        return radical
    else:
        return 0

def format(indice):
    """Renvoie le numéro mensuel au bon format
    Args:
        indice (int): indice mensuel au format nombre
    Returns:
        numero (str): indice mensuel au format chaine avec des 0
    """
    temp = str(indice)
    # Rajoute des zéros à gauche pour qu'il y ait 3 caractères
    numero = temp.zfill(3)
    return numero

def detection_annee(liste):
    """Indique l'année courante des factures de la liste
    Args:
        liste (str) : numéro de compte
    Returns:
        annee (str) : année la plus fréquente dans les numéros de facture
    """
    liste_annees = [extraire_annee(chaine[0]) for chaine in liste]
    compteur = Counter(liste_annees)
    # Récupérer uniquement la valeur la plus fréquente
    annee = str(compteur.most_common(1)[0][0])
    return annee

def extraction_factures(fichier, entite):
    """Liste les lignes client des écritures de vente
    Args:
        fichier (str) : chemin d'accès au journal des ventes
    Returns:
        listte_factures (list) : liste de mes factures comptabilisées
    """
    
     # Charger le fichier Excel
    wb = load_workbook(filename=fichier)

    # Accéder à la feuille du classeur
    feuille = wb.worksheets[entite]

    if feuille.cell(1, 1).value != 'Edition Journaux':
        print("Il ne s'agit pas d'un journal de vente")
        sys.exit()

    # Nombre de lignes dans la feuille
    nb_lignes = feuille.max_row
    liste_factures = []

    # Liste les factures clients
    # 1er élément : n° pièce
    # 2ème élément : jour
    # 3ème élément : libellé
    # 4ème élément : n° compte
    # 5ème élément : débit
    # 6ème élément : crédit
    for ligne in range(1, nb_lignes + 1):
        facture = []
        numCompte = feuille.cell(ligne, 4).value
        debit = feuille.cell(ligne, 6).value
        credit = feuille.cell(ligne, 7).value
        
        if compte_client_valide(numCompte, debit, credit):
            # Je ne fais pas de boucle pour prendre mes colonnes
            # dans l'ordre que je veux
            if feuille.cell(ligne, 2).value is not None:
                facture.append(feuille.cell(ligne, 2).value)
            else:
                facture.append('Numéro inconnu')
            facture.append(feuille.cell(ligne, 1).value)
            facture.append(feuille.cell(ligne, 3).value)
            facture.append(feuille.cell(ligne, 4).value)
            facture.append(feuille.cell(ligne, 6).value)
            facture.append(feuille.cell(ligne, 7).value)
            
            # N'ajoute dans la liste que les factures ayant un n° correct
            if extraire_indice(facture[0]) != 0:
                liste_factures.append(facture.copy())
    
    # Tri par numéro de pièce (1er élément) ma liste de factures
    # La key permet d'éviter que H et h ne soient pas triés uniformément
    liste_factures.sort(key=lambda x: x[0].upper())
            
    # Fermer le classeur Excel
    wb.close()
    
    return liste_factures

def creer_doublons(liste_factures):
    """Créer la liste des doublons de facture
    Args:
        liste_factures (list) : liste de mes factures de vente
    Returns:
        liste_doublons (list) : liste des doublons de facture
    """
    liste_doublons = []
    for index, facture in enumerate(liste_factures):
        if liste_factures[index][0] == liste_factures[index - 1][0]:
            liste_doublons.append(facture[0])
    return liste_doublons

def creer_manquants(liste_factures, liste_doublons):
    """Créer la liste des factures manquantes
    Args:
        liste_factures (list) : liste de mes factures de vente
        liste_doublons (list) : liste des doublons de facture
    Returns:
        liste_factures_manquantes (list) : liste des factures manquantes
    """
    
    liste_factures_manquantes = []

    # Je crée la liste des factures dont les numéros ne se suivent pas
    for index, facture in enumerate(liste_factures):
        numero_actuel = extraire_indice(liste_factures[index][0])
        numero_precedent = extraire_indice(liste_factures[index - 1][0])
        mois_actuel = extraire_mois(liste_factures[index][0])
        mois_precedent = extraire_mois(liste_factures[index - 1][0])
        annee_actuelle = extraire_annee(liste_factures[index][0])
        radical_actuel = extraire_radical(liste_factures[index][0])
        
        # Rajoute uniquement les numéros respectant ces conditions :
        # 1 : ne fait pas partie de la liste des doublons
        # 2 : l'année de la facture est bien l'année courante
        # 3 : le mois de la facture ne correspond pas à celui de la précédente
        # 4 : le numéro de la facture n'est pas le numéro 1
        if (facture[0] not in liste_doublons and
            str(annee_actuelle) == detection_annee(liste_factures) and
            mois_actuel != mois_precedent and
            numero_actuel != 1):
            for numero in range(1, numero_actuel):
                liste_factures_manquantes.append(
                        radical_actuel + format(numero))
        
        # Rajoute uniquement les numéros respectant ces conditions :
        # 1 : ne fait pas partie de la liste des doublons
        # 2 : l'année de la facture est bien l'année courante
        # 3 : le mois de la facture correspond à celui de la facture précédente
        # 4 : le numéro de la facture ne suit pas la précédente
        if (facture[0] not in liste_doublons and 
            str(annee_actuelle) == detection_annee(liste_factures) and 
            mois_actuel == mois_precedent and 
            numero_actuel != numero_precedent + 1):
                # Rajoute tous les numéros manquants entre les deux bornes
                for numero in range(numero_precedent + 1, numero_actuel):
                    liste_factures_manquantes.append(
                        radical_actuel + format(numero))
    
    return liste_factures_manquantes

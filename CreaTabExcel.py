import re
import sys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from CreaListes import extraire_annee, extraire_mois, detection_annee

"""
Crée le tableau excel de contrôle de la facturation et le met en forme

Auteur : Michaël AUGUSTE
"""

def creer_fichier_facturation(chemin):
    """Créer le fichier excel de facturation
    Args:
        chemin (str): nom du fichier
    Returns:
        chemin (str) : nouveau nom du fichier
    """
    
    while True:
        try:
            wb = load_workbook(filename=chemin)
            wb.close()  # Fermer le classeur après l'avoir ouvert
            chemin = incrementer_nom_fichier(chemin)
        except FileNotFoundError:
            print("Création du fichier de facturation")
            # Créer le fichier Excel
            wb = Workbook()
            wb.save(filename=chemin)
            wb.close()  # Fermer le classeur après l'avoir ouvert
            print(f"Le fichier {chemin} a été créé avec succès.")
            return chemin  # Retourne le nom du fichier une fois créé
        except Exception as e:
            print(f"Une erreur {e} s'est produite lors de l'ouverture du fichier")
            sys.exit  # Sortir de la boucle en cas d'autres erreurs

def incrementer_nom_fichier(chemin):
    """Incrémente le nom du fichier de (1)
    Args:
        chemin (str): nom du fichier
    Returns:
        nouveau_nom (str) : nouveau nom du fichier
    """
    
    # Sépare le nom du fichier et son extension
    nom_brut, extension = extraire_extension(chemin)
    
    # Chercher le motif de la forme "(x)" à la fin du nom de fichier
    match = re.search(r'(\(\d+\))$', nom_brut)
    
    if match:
        # Extraire le nombre entre parenthèses et l'incrémenter
        numero = int(match.group(1).strip('()')) + 1
        # Remplacer l'ancien numéro par le nouveau
        nouveau_nom = re.sub(r'\(\d+\)$', f'({numero})', nom_brut)
    else:
        # Ajouter "(1)" à la fin du nom du fichier
        nouveau_nom = nom_brut + " (1)"
    
    # Si aucune extension n'a été trouvée, je lui affecte l'extension suivante
    if extension == "":
        extension = ".xlsx"
        
    nouveau_nom = nouveau_nom + extension
    return nouveau_nom

def extraire_extension(chaine):
    """Sépare le nom du fichier de son extension
    Args:
        chaine (str) : nom du fichier
    Returns:
        nom_fichier_sans_extension (str) : nom brut du fichier
        extension (str) : extension du fichier
        chaine (str) : variable en paramètre si aucune extension n'est trouvée
    """
    
    resultat = re.search(r'(\.[^.]+)$', chaine)
    if resultat:
        extension = resultat.group(1)
        nom_fichier_sans_extension = chaine[:-len(extension)]
        return nom_fichier_sans_extension, extension
    else:
        return chaine, ""

def init_onglet_manquants(chemin):
    """Créer la structure de l'onglet "MANQUANTS"
    Args:
        chemin (str): chemin du fichier
    """
    
    # Charger le fichier Excel
    wb = load_workbook(filename=chemin)

    # Accéder à la feuille des factures manquantes
    feuille = wb.worksheets[0]
    feuille.title = "MANQUANTS"
    
    # Liste des noms des mois
    mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Créer un style nommé
    style_bleu_clair = NamedStyle(name="BleuClair")
    style_bleu_clair.fill = PatternFill(start_color="B4C6E7", 
                                        end_color="B4C6E7", 
                                        fill_type="solid")

    # Créer un style de bordure avec seulement la bordure supérieure
    border_top = Border(top=Side(style='thin'))

    # AZUR AUTO LA SEYNE
    # Créer le récap nombre de factures
    feuille.cell(1, 1).value = "TV + VVO"
    feuille.cell(1, 1).alignment = Alignment(horizontal='center')
    feuille.cell(1, 1).font = Font(bold=True, size=14)
    feuille.merge_cells('A1:D1')
    feuille.cell(3, 2).value = "Jrn ventes"
    feuille.cell(3, 2).alignment = Alignment(horizontal='center')
    feuille.cell(3, 2).font = Font(bold=True)
    feuille.cell(3, 3).value = "Fact. Clé USB"
    feuille.cell(3, 3).alignment = Alignment(horizontal='center')
    feuille.cell(3, 3).font = Font(bold=True)
    feuille.cell(3, 4).value = "Ecart"
    feuille.cell(3, 4).alignment = Alignment(horizontal='center')
    feuille.cell(3, 4).font = Font(bold=True)
    # Créer les lignes mensuelles du récap
    for i, nom_mois in enumerate(mois):
        num = i + 4
        feuille.cell(row=num, column=1, value=nom_mois)
        feuille.cell(row=num, column=3).style = style_bleu_clair
        formule = "=C" + str(num) + "-B" + str(num)
        feuille.cell(row=num, column=4, value=formule)
        feuille.cell(num, 2).alignment = Alignment(horizontal='center')
        feuille.cell(num, 3).alignment = Alignment(horizontal='center')
        feuille.cell(num, 4).alignment = Alignment(horizontal='center')
    # Créer les totaux du récap
    for i in range(0, 3):
        num = i + 2
        col = get_column_letter(num)
        feuille.cell(16, num).value = "_______"
        feuille.cell(16, num).alignment = Alignment(horizontal='center')
        formule = "=SUM(" + col + "4:" + col + "16)"
        feuille.cell(17, num).value = formule
        feuille.cell(17, num).alignment = Alignment(horizontal='center')
        feuille.cell(17, num).font = Font(bold=True)
    feuille.cell(17, 1).value = "nb factures"
    feuille.cell(18, 1).value = "verif lignes"
    feuille.cell(18, 2).alignment = Alignment(horizontal='center')
    
    # AZUR AUTO HYERES
    # Créer le récap nombre de factures
    feuille.cell(1, 6).value = "TVH"
    feuille.cell(1, 6).alignment = Alignment(horizontal='center')
    feuille.cell(1, 6).font = Font(bold=True, size=14)
    feuille.merge_cells('F1:I1')
    feuille.cell(3, 7).value = "Jrn ventes"
    feuille.cell(3, 7).alignment = Alignment(horizontal='center')
    feuille.cell(3, 7).font = Font(bold=True)
    feuille.cell(3, 8).value = "Fact. Clé USB"
    feuille.cell(3, 8).alignment = Alignment(horizontal='center')
    feuille.cell(3, 8).font = Font(bold=True)
    feuille.cell(3, 9).value = "Ecart"
    feuille.cell(3, 9).alignment = Alignment(horizontal='center')
    feuille.cell(3, 9).font = Font(bold=True)
    # Créer les lignes mensuelles du récap
    for i, nom_mois in enumerate(mois):
        num = i + 4
        feuille.cell(row=num, column=6, value=nom_mois)
        feuille.cell(row=num, column=8).style = style_bleu_clair
        formule = "=H" + str(num) + "-G" + str(num)
        feuille.cell(row=num, column=9, value=formule)
        feuille.cell(num, 7).alignment = Alignment(horizontal='center')
        feuille.cell(num, 8).alignment = Alignment(horizontal='center')
        feuille.cell(num, 9).alignment = Alignment(horizontal='center')
    # Créer les totaux du récap
    for i in range(0, 3):
        num = i + 7
        col = get_column_letter(num)
        feuille.cell(16, num).value = "_______"
        feuille.cell(16, num).alignment = Alignment(horizontal='center')
        formule = "=SUM(" + col + "4:" + col + "16)"
        feuille.cell(17, num).value = formule
        feuille.cell(17, num).alignment = Alignment(horizontal='center')
        feuille.cell(17, num).font = Font(bold=True)
    feuille.cell(17, 6).value = "nb factures"
    feuille.cell(17, 10).value = "nb factures manquantes"
    feuille.cell(18, 6).value = "verif lignes"
    feuille.cell(18, 7).alignment = Alignment(horizontal='center')
    
    # Créer le début du tableau des factures manquantes
    feuille.cell(20, 1).border = border_top
    feuille.cell(20, 1).alignment = Alignment(horizontal='center')
    feuille.cell(20, 1).font = Font(bold=True)
    for i, nom_mois in enumerate(mois):
        num = i + 2
        feuille.cell(row=20, column=num, value=nom_mois)
        feuille.cell(20, num).border = border_top
        feuille.cell(20, num).alignment = Alignment(horizontal='center')
        feuille.cell(20, num).font = Font(bold=True)
    
    # Mettre un commentaire pour faire attention
    commentaire = ("Attention, il faut bien vérifier que le dernier numéro " +
                   "de facture de chaque mois ne soit pas manquant")
    feuille.cell(1, 11).value = commentaire
    commentaire = "car ils ne peuvent pas être détectés automatiquement"
    feuille.cell(2, 11).value = commentaire
    
    # Ajuster à 12 la taille des 13 premières colonnes
    for col_num in range(1, 14):
        column_letter = get_column_letter(col_num)
        feuille.column_dimensions[column_letter].width = 12
    
    wb.save(chemin)
    wb.close()
    
def onglet_factures(liste_factures, entite, chemin):
    """Liste les factures de chaque entité dans des onglets
    Args:
        liste_factures (list): liste des factures
        entite (int) : code 0 pour la Seyne et 1 pour Hyères
        chemin (str) : chemin du fichier
    """
    
    # Charger le fichier Excel
    wb = load_workbook(filename=chemin)
    
    annee_globale = detection_annee(liste_factures)

    # Créer la feuille listant les factures de l'entité
    if entite == 0:
        nom_onglet = "TV+VVO " + annee_globale
    elif entite == 1:
        nom_onglet = "TVH " + annee_globale
    else:
        nom_onglet = "inconnu"
    
    feuille = wb.create_sheet(nom_onglet)
    
    # Créer un style de bordure avec seulement la bordure supérieure
    border_top = Border(top=Side(style='thin'))
    
    # Nommer les en-têtes
    feuille.cell(1, 1).value = "Jour"
    feuille.cell(1, 2).value = "Pièce"
    feuille.cell(1, 3).value = "Libellé de l'écriture"
    feuille.cell(1, 4).value = "Compte"
    feuille.cell(1, 5).value = "Débit"
    feuille.cell(1, 6).value = "Crédit"
    feuille.cell(1, 8).value = "Index"
    feuille.cell(1, 9).value = "N° Réel"
    feuille.cell(1, 10).value = "Ecart"
    
    # Mettre les en-têtes en gras et centré
    for i in range(1, 11):
        feuille.cell(1, i).font = Font(bold=True)
        feuille.cell(1, i).alignment = Alignment(horizontal='center')
    
    # Ajuster la taille des colonnes
    feuille.column_dimensions['A'].width = 4
    feuille.column_dimensions['B'].width = 13
    feuille.column_dimensions['C'].width = 30
    feuille.column_dimensions['D'].width = 11
    feuille.column_dimensions['E'].width = 11
    feuille.column_dimensions['F'].width = 11
    feuille.column_dimensions['H'].width = 6
    feuille.column_dimensions['I'].width = 7
    feuille.column_dimensions['J'].width = 5
    
    # Alimenter le tableau avec les factures de l'entité
    # Je démarre l'index à 2 pour qu'il corresponde au numéro de ligne
    for index, facture in enumerate(liste_factures, start=2):
        feuille.cell(index, 1).value = facture[1]
        feuille.cell(index, 2).value = facture[0]
        feuille.cell(index, 3).value = facture[2]
        feuille.cell(index, 4).value = facture[3]
        feuille.cell(index, 5).value = facture[4]
        feuille.cell(index, 6).value = facture[5]
        feuille.cell(index, 1).alignment = Alignment(horizontal='center')
        feuille.cell(index, 2).alignment = Alignment(horizontal='center')
        feuille.cell(index, 4).alignment = Alignment(horizontal='center')
        feuille.cell(index, 5).alignment = Alignment(horizontal='center')
        feuille.cell(index, 6).alignment = Alignment(horizontal='center')
        
        # Je donne une valeur à la colonne Index
        # Je recalcule l'index car il démarre à 2 au lieu de 0
        mois_actuel = extraire_mois(liste_factures[index - 2][0])
        mois_precedent = extraire_mois(liste_factures[index - 3][0])
        annee_actuelle = extraire_annee(facture[0])
        if annee_actuelle == int(annee_globale):
            if mois_actuel != mois_precedent:
                feuille.cell(index, 8).value = 1
                
                # Si on change de mois de facturation, insère une bordure
                for col in range(1, 11):
                    feuille.cell(index, col).border = border_top
                
            else:
                formule = "=RIGHT(B" + str((index - 1)) + ",3)+1"
                feuille.cell(index, 8).value = formule
            feuille.cell(index, 8).alignment = Alignment(horizontal='center')
                
            # Je donne une valeur à la colonne N° Réel
            formule = "=RIGHT(B" + str(index) + ",3)+0"
            feuille.cell(index, 9).value = formule
            feuille.cell(index, 9).alignment = Alignment(horizontal='center')
            
            # Je donne une valeur à la colonne Ecart
            formule = "=IF(H" + str(index) + "=I" + str(index) + ",0,1)"
            feuille.cell(index, 10).value = formule
            feuille.cell(index, 10).alignment = Alignment(horizontal='center')
    
    # Appliquer le filtre automatique
    feuille.auto_filter.ref = feuille.dimensions
    
    # J'insère mes nombres récapitulatifs de factures mensuelles
    nb_recap_factures(wb, feuille, int(annee_globale))
    
    wb.save(chemin)
    wb.close()

def numero_mois(nom_mois):
    """Recherche le numéro du mois à partir de son nom
    Args:
        nom_mois (str): nom du mois de l'année
    Returns:
        date.month (int): numéro du mois de l'année
    """
    mois = {
        'Janvier': 1,
        'Février': 2,
        'Mars': 3,
        'Avril': 4,
        'Mai': 5,
        'Juin': 6,
        'Juillet': 7,
        'Août': 8,
        'Septembre': 9,
        'Octobre': 10,
        'Novembre': 11,
        'Décembre': 12
    }
    return mois.get(nom_mois)

def nb_recap_factures(wb, feuille, annee_globale):
    """Permet d'insérer les nombres totaux de factures par mois
    sur le 1er onglet "MANQUANTS"
    Args:
        wb (workbook) : fichier excel sur lequel on travaille
        feuille (worksheet) : feuille de l'entité
        annee_globale (int) : année de facturation
    """
    
    # Définition de l'onglet comportant le récapitulatif à alimenter
    feuilleRecap = wb.worksheets[0]
    
    # Tableau qui va comporter les cellules début et fin de chaque mois
    tab_recap_mensuel = []
    tab_mensuel = []
    
    # Affecter les lignes de début et de fin de chaque mois
    debutCell = None
    finCell = None
    
    for index in range(1, feuille.max_row + 1):
        mois_actuel = extraire_mois(feuille.cell(index, 2).value)
        mois_suivant = extraire_mois(feuille.cell(index + 1, 2).value)
        annee_actuelle = extraire_annee(feuille.cell(index, 2).value)
        
        # J'affecte debutCell si c'est la 1ère occurence du mois
        # Ce sera faux pour janvier si c'est une facture N-1 de janvier
        if annee_actuelle == annee_globale and debutCell is None:
            debutCell = feuille.cell(index, 2).coordinate

        # J'affecte finCell si c'est la dernière occurence du mois
        if mois_suivant != mois_actuel and debutCell is not None:
            finCell = feuille.cell(index, 2).coordinate
            tab_mensuel.append(mois_actuel)
            tab_mensuel.append(debutCell)
            tab_mensuel.append(finCell)
            tab_recap_mensuel.append(tab_mensuel.copy())
            tab_mensuel.clear()
            debutCell = None
            finCell = None
    
    # Déterminer la colonne à partir de l'index de la feuille dans le classeur
    index_feuille = wb.sheetnames.index(feuille.title)
    if index_feuille == 1:
        col = 1
    elif index_feuille == 2:
        col = 6
    else:
        print("impossible de trouver l'entité" +
              "à partir de la position de la feuille")
        return
    
    # J'inscrit le nombre de factures mensuelles sur les lignes entre 1 et 20
    for index in range(1, 20):
        num = numero_mois(feuilleRecap.cell(index, col).value)
        for mois in tab_recap_mensuel:
            if num == mois[0]:
                formule = ("=COUNTA('" + feuille.title + "'!" + 
                           mois[1] + ":" + mois[2] + ")")
                feuilleRecap.cell(index, col + 1).value = formule
        
        # J'inscrit le nombre total de factures dans verif lignes
        if feuilleRecap.cell(index, col).value == "verif lignes":
            formule = ("=COUNTA('" + feuille.title + "'!" + 
                           tab_recap_mensuel[0][1] + ":" + 
                           tab_recap_mensuel[-1][2] + ")")
            feuilleRecap.cell(index, col + 1).value = formule
    
def onglet_manquants(liste_manquants, liste_doublons, entite, chemin):
    """Affiche la liste des factures manquantes et des doublons

    Args:
        liste_manquants (list): liste des factures manquantes
        liste_doublons (list): liste des doublons de facture
        entite (int): numéro de l'entité
        chemin (str): chemin d'accès au fichier Excel
    """
    # Charger le fichier Excel
    wb = load_workbook(filename=chemin)

    # Créer la feuille listant les factures de l'entité
    if entite == 0:
        nom_entite = "TV + VVO"
    elif entite == 1:
        nom_entite = "TVH"
    else:
        nom_entite = "inconnu"
    
    feuille = wb.worksheets[0]
    
    # Créer un style de bordure avec seulement la bordure supérieure
    border_top = Border(top=Side(style='thin'))
    
    # Nombre de lignes dans la feuille
    nb_lignes = feuille.max_row
    
    # Définir la ligne sur laquelle j'ai la liste des mois
    for ligne in range(1, nb_lignes + 1):
        for col in range(2, 14):
            num = numero_mois(feuille.cell(ligne, col).value)
            if num == col - 1:
                ligne_entete = ligne
                break
    
    # J'insère ma bordure supérieure
    for col in range(1, 14):
        feuille.cell(nb_lignes + 1, col).border = border_top
    
    # Je met un titre à mes factures manquantes
    feuille.cell(nb_lignes + 1, 1).value = nom_entite
    feuille.cell(nb_lignes + 2, 1).value = "Factures"
    feuille.cell(nb_lignes + 3, 1).value = "manq."
    for index in range(nb_lignes + 1, nb_lignes + 4):
        feuille.cell(index, 1).alignment = Alignment(horizontal='center')
        feuille.cell(index, 1).font = Font(bold=True)
    
    # J'insère mes numéros de factures manquantes
    for col in range(2, 14):
        # Je crée ma variable increment pour augmenter de ligne
        # au fur et à mesure que je renseigne mes numéros de facture
        increment = 2
        
        num = numero_mois(feuille.cell(ligne_entete, col).value)
        
        for manquant in liste_manquants:
            if num == extraire_mois(manquant):
                feuille.cell(nb_lignes + increment, col).value = manquant
                feuille.cell(nb_lignes + increment, col).alignment = Alignment(
                    horizontal='center')
                increment += 1
    
    # Je réactualise mon nombre de lignes dans la feuille
    nb_lignes = feuille.max_row
    
    # Je met un titre à mes doublons
    feuille.cell(nb_lignes + 2, 1).value = "Doublons"
    feuille.cell(nb_lignes + 2, 1).alignment = Alignment(
        horizontal='center')
    feuille.cell(nb_lignes + 2, 1).font = Font(bold=True)
    
    # J'insère mes numéros de doublons de facture
    for col in range(2, 14):
        # Je crée ma variable increment pour augmenter de ligne
        # au fur et à mesure que je renseigne mes numéros de facture
        increment = 2
        
        num = numero_mois(feuille.cell(ligne_entete, col).value)
        
        for doublon in liste_doublons:
            if num == extraire_mois(doublon):
                feuille.cell(nb_lignes + increment, col).value = doublon
                feuille.cell(nb_lignes + increment, col).alignment = Alignment(
                    horizontal='center')
                increment += 1
    
    wb.save(chemin)
    wb.close()
            
        
